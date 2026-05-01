"""
Excel Writer Module

Writes scraped funda property data to an Excel (.xlsx) file
using openpyxl.  Each property becomes one row with columns:

  URL | Images | Energielabel | Description |
  Asking Price | Bidding Price |
  Agency Name | Agency Phone | Agency Email | Agency Website
"""
from pathlib import Path
from typing import List, Dict, Any
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from ..utils.logger import setup_logger

logger = setup_logger('funda.excel')

# Column definitions: (header, dict_key, width)
COLUMNS = [
    # Identity
    ('Property URL',        'url',                  50),
    ('Address',             'address',              35),
    ('Listed Since',        'listed_since',         14),
    ('Days on Market',      'days_on_market',       14),
    # Pricing
    ('Asking Price (€)',    'asking_price',         18),
    ('Price / m² (€)',      'price_per_m2',         14),
    # Dimensions
    ('Living Area (m²)',    'living_area',          14),
    ('Plot Area (m²)',      'plot_area',            14),
    ('Volume (m³)',         'volume',               12),
    ('Rooms',               'rooms',                10),
    ('Bedrooms',            'bedrooms',             10),
    ('Floors',              'floors',               10),
    # Property identity
    ('Property Type',       'property_type',        22),
    ('Build Type',          'build_type',           18),
    ('Construction Year',   'construction_year',    16),
    ('Roof Type',           'roof_type',            18),
    # Energy & condition
    ('Energy Label',        'energielabel',         13),
    ('Heating',             'heating',              25),
    ('Insulation',          'insulation',           35),
    ('Maintenance Inside',  'maintenance_inside',   18),
    ('Maintenance Outside', 'maintenance_outside',  18),
    # Outdoor & parking
    ('Garden',              'garden',               18),
    ('Garden Orientation',  'garden_orientation',   18),
    ('Parking',             'parking',              22),
    # Financial / legal
    ('VVE (€/month)',       'vve_contribution',     16),
    ('Erfpacht',            'erfpacht',             12),
    ('Acceptance',          'acceptance',           16),
    # Description & media
    ('Description',         'description',          80),
    ('Images',              '_images_joined',       80),
    # Agency
    ('Agency Name',         'agency_name',          30),
    ('Agency Phone',        'agency_phone',         20),
    ('Agency Email',        'agency_email',         30),
    ('Agency Website',      'agency_website',       40),
]

# Columns that should be formatted as integers
INT_COLUMNS = {
    'asking_price', 'price_per_m2',
    'living_area', 'plot_area', 'volume',
    'rooms', 'bedrooms', 'construction_year', 'days_on_market',
}

# Styles
HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
CELL_ALIGN = Alignment(vertical='top', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)
PRICE_FORMAT = '#,##0'


class ExcelWriter:
    """Writes property data to an Excel workbook."""

    def __init__(self, output_dir: Path, filename_prefix: str = 'funda_properties'):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.filename_prefix = filename_prefix

    def write(self, properties: List[dict], filename: str = None) -> Path:
        """
        Write property data to an Excel file.

        Args:
            properties: list of property dicts (from PropertyScraper)
            filename: optional override for the output filename

        Returns:
            Path to the created Excel file
        """
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{self.filename_prefix}_{timestamp}.xlsx"

        filepath = self.output_dir / filename

        wb = Workbook()
        ws = wb.active
        ws.title = 'Properties'

        # ── Header row ────────────────────────────────────────
        for col_idx, (header, _, width) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze top row
        ws.freeze_panes = 'A2'

        # ── Data rows ─────────────────────────────────────────
        for row_idx, prop in enumerate(properties, 2):
            # Pre-process: join photo URLs with comma
            prop['_images_joined'] = ', '.join(prop.get('photo_urls', []))

            for col_idx, (_, key, _) in enumerate(COLUMNS, 1):
                value = prop.get(key, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = CELL_ALIGN
                cell.border = THIN_BORDER

                # Format numeric columns
                if key in INT_COLUMNS and isinstance(value, (int, float)) and value:
                    cell.number_format = PRICE_FORMAT

        # ── Auto-filter ───────────────────────────────────────
        if properties:
            last_col = get_column_letter(len(COLUMNS))
            last_row = len(properties) + 1
            ws.auto_filter.ref = f"A1:{last_col}{last_row}"

        # ── Summary sheet ─────────────────────────────────────
        self._add_summary_sheet(wb, properties)

        # ── Save ──────────────────────────────────────────────
        wb.save(filepath)
        logger.info(f"  ✓ Excel written: {filepath}")
        logger.info(f"    {len(properties)} properties, {len(COLUMNS)} columns")

        return filepath

    def _add_summary_sheet(self, wb: Workbook, properties: List[dict]) -> None:
        """Add a summary sheet with aggregate statistics."""
        ws = wb.create_sheet('Summary')

        if not properties:
            ws.cell(row=1, column=1, value='No properties scraped')
            return

        prices   = [p['asking_price'] for p in properties if p.get('asking_price')]
        ppm2     = [p['price_per_m2'] for p in properties if p.get('price_per_m2')]
        dom_vals = [p['days_on_market'] for p in properties if p.get('days_on_market')]
        areas    = [p['living_area'] for p in properties if p.get('living_area')]

        def avg(lst): return sum(lst) / len(lst) if lst else 0

        stats = [
            ('Funda Property Scraper — Summary', ''),
            ('', ''),
            ('Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ('Total Properties', len(properties)),
            ('', ''),
            ('Price Statistics', ''),
            ('Min Asking Price',  f"€ {min(prices):,.0f}"     if prices else 'N/A'),
            ('Max Asking Price',  f"€ {max(prices):,.0f}"     if prices else 'N/A'),
            ('Avg Asking Price',  f"€ {avg(prices):,.0f}"     if prices else 'N/A'),
            ('', ''),
            ('Price per m² Stats', ''),
            ('Min Price/m²',      f"€ {min(ppm2):,.0f}"       if ppm2 else 'N/A'),
            ('Max Price/m²',      f"€ {max(ppm2):,.0f}"       if ppm2 else 'N/A'),
            ('Avg Price/m²',      f"€ {avg(ppm2):,.0f}"       if ppm2 else 'N/A'),
            ('', ''),
            ('Market Stats', ''),
            ('Avg Living Area',   f"{avg(areas):,.0f} m²"     if areas else 'N/A'),
            ('Avg Days on Market',f"{avg(dom_vals):,.1f} days" if dom_vals else 'N/A'),
            ('Fresh (<7 days)',   sum(1 for d in dom_vals if d < 7)),
            ('Stale (>21 days)',  sum(1 for d in dom_vals if d > 21)),
            ('', ''),
            ('Energy Labels', ''),
        ]

        # Count energy labels
        labels = {}
        for p in properties:
            label = p.get('energielabel', 'Unknown')
            labels[label] = labels.get(label, 0) + 1
        for label in sorted(labels.keys()):
            stats.append((f'  {label}', labels[label]))

        for row_idx, (label, value) in enumerate(stats, 1):
            ws.cell(row=row_idx, column=1, value=label).font = Font(
                bold=row_idx in (1, 6, 11, 15)
            )
            ws.cell(row=row_idx, column=2, value=value)

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
